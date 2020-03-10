"""Collect the number of advertisements from the following categories,
    - tvs
    - mobile-phones
    - computers-tablets
    - air-conditions-electrical-fittings
    - electronic-home-appliances(from selected item from electric_home_appliance)
"""

import re
import requests
import time
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from openpyxl import Workbook

full_dict = {}
serp_url = 'https://ikman.lk/data/serp'
ad_url = 'https://ikman.lk/en/ad/'
tv_brands = 'samsung|sony|innovex|lg|panasonic|haier|toshiba|jvc|philips|' \
            'sharp|nec|sansui|hisense|abans'
mobile_brands = 'apple|samsung|huawei|xiaomi|nokia|oppo|sony|vivo|lg|htc|' \
                'oneplus|lenovo|microsoft|micromax|google|e-tel|blackberry|' \
                'china-mobile|greentel|asus|dialog|zigo|sony-ericsson|' \
                'motorola|sky|acer|alcatel|zte|ag-tel|ipro|q-mobile|palm|' \
                'i-mate|realme|panasonic|doogee|sharp|umidigi|ccit|energizer|' \
                'intex|docomo'
pc_brands = 'intel|hp|dell|samsung|asus|lenovo|acer|toshiba|microsoft|sony|' \
            'ibm|huawei'
electric_home_appliance = {
    'refrigerator': 'ref|free|fri',
    'washing_machine': 'wash',
    'microwave': 'micro|wave',
    'rice_cooker': 'rice-cooker|rice-steamer',
    'gas_cooker': 'gas-cooker|gas-stove',
    'kettle': 'kettle',
    'blender': 'blender',
    'oven': 'oven'
}

date_format = '%Y-%m-%d'
days_re_search = 'day'
bump_re_search = 'bump_up'
min_hour_re_search = 'minutes|minute|hours|hour|just now'


def validate_day_and_bump_up(timestamp: str, search: str) -> bool:
    """To validate both days and bump_up formats.

    Args:
        timestamp: String type timestamp.
        search: Search value for regex.

    Returns:
        The return value. True for success, False otherwise.
    """

    _search = re.search(search, timestamp)

    if _search and search == days_re_search:
        return True
    elif _search and search == bump_re_search:
        return True
    elif _search and search == min_hour_re_search:
        return True
    return False


def validate_day(timestamp: str, datetime_obj: datetime) -> int:
    """This function helps for maintaining 3 types such as break, continue and
    pass inside the called function.

    Args:
        timestamp: String type timestamp.

    Returns:
        An integer value.

            0: pass
            1: break
            2: continue
    """

    timestamp_days = int(timestamp.split(' ')[0])
    days = (datetime.now() - datetime_obj).days
    if timestamp_days == days:
        return 0
    elif timestamp_days > days:
        return 1
    elif timestamp_days < days:
        return 2


def timedelta_func(**kwargs) -> int:
    """Subtract hours or minutes from datetime object.

    Args:
        **kwargs: Arbitrary keyword arguments.

    Returns:
        An integer value.

            0: pass
            1: break
            2: continue
    """

    entered_date = kwargs.pop('datetime_obj')
    datetime_obj = datetime.now() - timedelta(**kwargs)
    print(f'timedelta_func: {datetime_obj.date()} | {entered_date.date()}')
    if datetime_obj.date() == entered_date.date():
        return 0
    elif datetime_obj.date() < entered_date.date():
        return 1
    elif datetime_obj.date() > entered_date.date():
        return 2


def validate_hours_minutes(timestamp, datetime_obj) -> int:
    """Validate when the timestamp with hours or minutes

    Args:
        timestamp: String type timestamp.

    Returns:
        An integer value.

            0: pass
            1: break
            2: continue
    """

    # days = (datetime.now() - datetime_obj).days
    if timestamp == 'just now' and datetime.now().date() == datetime_obj.date():
        return 0
    elif timestamp == 'just now' and datetime.now().date() > datetime_obj.date():
        return 1
    elif timestamp == 'just now' and datetime.now().date() < datetime_obj.date():
        return 2

    try:
        split_timestamp = int(timestamp.split(' ')[0])

        _search = re.search(min_hour_re_search, timestamp)
        if not _search:
            return 2

        group = _search.group()
        if group in ['minutes', 'minute']:
            return timedelta_func(minutes=split_timestamp, datetime_obj=datetime_obj)
        elif group in ['hours', 'hour']:
            return timedelta_func(hours=split_timestamp, datetime_obj=datetime_obj)
        return 2
    except Exception as e:
        return 2


def get_query_params(sort='date', order='desc', category_slug='tvs', location_slug='sri-lanka', page=1, category_value=870) -> str:
    """Create the query params value for `serp_url`.

    Args:
        category_value: Integer type category of the category_slug.
        sort: The sort value such as date or price.
        order: Order value should be desc or asc.
        category_slug: The category slug value.
        location_slug: The location slug
        page: Current page.

    Returns:
        The completed query params value.
    """

    sort = f'sort={sort}'
    order = f'order={order}'
    category_slug = f'categorySlug={category_slug}'
    location_slug = f'locationSlug={location_slug}'
    page = f'page={page}'
    category_value = f'category={category_value}'

    return f'?{sort}&{order}&{category_slug}&{location_slug}&{page}&{category_value}'


def get_ads(category_slug: str, page: int, category_value: int) -> list:
    """Get the list of ads from `serp_url`.

    Args:
        category_slug: The category slug to query.
        page: Current page value.

    Returns:
        The list of ads.
    """
    while True:
        query_params = get_query_params(category_slug=category_slug, page=page, category_value=category_value)
        print(f'query_params: {query_params}')
        time.sleep(3)
        response = requests.get(f'{serp_url}{query_params}')
        if response.status_code != 200:
            time.sleep(5)
            continue

        print(f'response: {response}')
        # print(response.json())
        ads: list = response.json()['ads']
        return ads


def base_logic(timestamp):
    print('#' * 100)
    print(f'timestamp: {timestamp}')
    if validate_day_and_bump_up(timestamp, bump_re_search):
        print(f'validate_day_and_bump_up: 1')
        return 2

    if validate_day_and_bump_up(timestamp, days_re_search):
        print(f'validate_day_and_bump_up: 2')
        validated = validate_day(timestamp, datetime_obj)
        if validated == 1:
            print(f'validate_day: 1')
            return 1
        elif validated == 2:
            print(f'validate_day: 2')
            return 2

    if validate_day_and_bump_up(timestamp, min_hour_re_search):
        print(f'validate_day_and_bump_up: 3')
        validated = validate_hours_minutes(timestamp, datetime_obj)
        if validated == 1:
            print(f'validate_day: 1')
            return 1
        elif validated == 2:
            print(f'validate_day: 2')
            return 2


def calculate_common(brands, category='tvs', brand_position=2, filtering=True, datetime_obj=None, page=None, category_value=None) -> dict and int:
    """This function handled following categories commonly.
        - tvs
        - mobile-phones
        - computers-tablets
        - air-conditions-electrical-fittings

    Args:
        brands: Brand names which can be easily grab from category slug field.
        category: One of the category value listed above.
        brand_position: The positions Where we can find Brand.
        filtering: False mean just get the count without taking brand name.

    Returns:
        Number of collected brands with the dict.
    """

    data_dict = {}

    while True:
        ads: list = get_ads(category, page, category_value)
        status = True

        for i in ads:
            slug: str = i['slug']
            timestamp: str = i['timeStamp']

            validated = base_logic(timestamp)
            if validated == 1:
                status = False
                break
            elif validated == 2:
                continue

            print('*' * 100)
            print(f'datetime_obj: {datetime_obj}\ntimestamp: {timestamp}\nslug: {slug}')

            brand_search = re.search(brands, slug)
            if not filtering:
                brand = 'no_brand'
                print('no_brand')
            elif brand_search:
                print(f'group: {brand_search.group()}')
                brand = brand_search.group()
            else:
                url = f'{ad_url}{slug}'
                time.sleep(3)
                content = requests.get(url).text
                soup = BeautifulSoup(content, 'html.parser')
                dd = soup.find_all('dd')
                if dd:
                    brand = dd[brand_position-1].text
                else:
                    brand = 'Other Brand'
                brand: str = brand.lower().replace(' ', '_')
                print(f'brand: {brand}')

            if brand not in data_dict.keys():
                data_dict[brand] = 0

            data_dict[brand] += 1
            print(f'category: {brand} added to the database')

        if not status:
            return data_dict, page

        page += 1


def recurse_home_appliance(content: str) -> str:
    """Check whether the value of `electric_home_appliance` in the `content`. If
        this is match any of the value in `electric_home_appliance` mean, we are
        expecting to count that category which is match the key of that value.

    Args:
        content: Slug value from the ads.

    Returns:
        Returns the matches key to continue the calculated category.
    """

    for key, value in electric_home_appliance.items():
        search = re.search(value, content)
        if search:
            return key


def calculate_home_appliance(category='electronic-home-appliances', datetime_obj=None, page=None, category_value=None) -> dict and int:
    """To handle the electronic-home-appliances category from the `serp_url`.

    Args:
        category: The category field.

    Returns:
        Number of collected brands with the dict.
    """

    data_dict = {}

    while True:
        ads: list = get_ads(category, page, category_value)
        status = True

        for i in ads:
            slug = i['slug']
            timestamp = i['timeStamp']
            category = recurse_home_appliance(slug)

            print('*' * 100)
            print(f'datetime_obj: {datetime_obj}\ntimestamp: {timestamp}\nslug: {slug}\ncategory: {category}')

            if not category:
                print(f'not category: {category}')
                continue

            validated = base_logic(timestamp)
            if validated == 1:
                status = False
                break
            elif validated == 2:
                continue

            if category not in data_dict.keys():
                data_dict[category] = 0

            data_dict[category] += 1
            print(f'category: {category} added to the database')

        if not status:
            return data_dict, page

        page += 1


def validate_date(str_date: str) -> datetime or None:
    """Check whether the given date format is match to YYYY-MM-DD.

    Args:
        str_date: String type date.

    Returns:
        If given date beyond the current date or the format of the given date
        doesn't match with YYYY-MM-DD return None and otherwise returns datetime
        object.
    """

    try:
        # global entered_date
        entered_date = datetime.strptime(str_date, date_format)
        num_of_days = (datetime.now() - entered_date).days

        if num_of_days > 60:
            print('cannot display more than 60 days history')
            return None
        elif not entered_date > datetime.now():
            return entered_date

        print('you entered a future date')
    except ValueError as e:
        print('enter your date with following YYYY-MM-DD format')
    except Exception as e:
        print(e)
    return None


def validate_day_count(count: str) -> int or None:
    """Convert str into int and check the validations.

    Args:
        count: String type number of days.

    Returns:
        If validation pass will return an int type object, otherwise None.
    """
    try:
        int_count = int(count)
        if int_count > 59 or int_count < 0:
            print('valid day count should be in 0-59')
            return None
        return int_count
    except ValueError as e:
        print('enter valid integer upto 59')
    except Exception as e:
        print(e)
    return None


if __name__ == "__main__":
    while True:
        num_of_days = input('\nEnter number of days(0 for today, 1 for yesterday...): ')
        date_count = validate_day_count(num_of_days)
        if isinstance(date_count, int):
            break

    tv_page = 1
    mobile_page = 1
    pc_tablets_page = 1
    ac_page = 1
    home_appliance_page = 1

    for i in range(date_count+1):
        datetime_obj = datetime.now() - timedelta(days=i)

        # calculate tvs
        category = 'tvs'
        tv, tv_page = calculate_common(tv_brands, category, 2, datetime_obj=datetime_obj, page=tv_page, category_value=865)
        full_dict[category] = tv

        # calculate mobile phones
        category = 'mobile-phones'
        mobile, mobile_page = calculate_common(mobile_brands, category, 2, datetime_obj=datetime_obj, page=mobile_page, category_value=429)
        full_dict[category] = mobile

        # calculate computers and tablets
        category = 'computers-tablets'
        pc_tablets, pc_tablets_page = calculate_common(pc_brands, category, 3, datetime_obj=datetime_obj, page=pc_tablets_page, category_value=898)
        apple_macbook = pc_tablets.pop('apple_macbook', 0)
        apple_ipad = pc_tablets.pop('apple_ipad', 0)
        apple_imac = pc_tablets.pop('apple_imac', 0)
        apple_count = apple_macbook + apple_ipad + apple_imac
        pc_tablets['apple'] = apple_count
        full_dict[category] = pc_tablets

        # calculate air conditions electrical fittings
        category = 'air-conditions-electrical-fittings'
        ac, ac_page = calculate_common('', category, filtering=False, datetime_obj=datetime_obj, page=ac_page, category_value=871)
        full_dict[category] = ac

        # calculate electronic home appliances
        category = 'electronic-home-appliances'
        home_appliance, home_appliance_page = calculate_home_appliance(category, datetime_obj=datetime_obj, page=home_appliance_page, category_value=870)
        full_dict[category] = home_appliance

        print(f'full_dict: {full_dict}')

        workbook = Workbook()
        sheet = workbook.active

        x_count = 1
        for key, value in full_dict.items():
            sheet.cell(row=x_count, column=1, value=key)
            for key_1, value_1 in value.items():
                sheet.cell(row=x_count, column=2, value=key_1)
                sheet.cell(row=x_count, column=3, value=value_1)
                x_count += 1
            x_count += 1

        datetime_now = datetime_obj.strftime(date_format)
        workbook.save(filename=f"{datetime_now}.xlsx")
